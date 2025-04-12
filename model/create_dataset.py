from openpyxl import Workbook
from openpyxl import load_workbook
import json

def empty(curr_sheet, curr_row):
    for i in range(0, 3):
        for j in range(1, 10):
            if curr_sheet.cell(row=curr_row+i, column=j).value != None:
                return False

    return True

wb = load_workbook('./data/dataset.xlsx')

curr_sheet = wb['Таблица']

i = 7
data = []
txt_user = "Проанализируй сообщение и распредели по группам\n Вот пример сообщения: "
txt_system = "Ты — помощник агронома, который должен распределить информацию из собщения по группам: Дата, Подразделение, Операция, Культура, За день га, С начала операции га,Вал за день ц, Вал с начала ц"
txt_answer = "\n"
col_names = ['Дата', 'Подразделение', 'Операция', 'Культура', 'За день, га', 'С начала операции, га', 'Вал за день, ц', 'Вал с начала, ц']

while not(empty(curr_sheet, i)):
    if curr_sheet.cell(row=i, column=1).value != None:
        if curr_sheet.cell(row=i, column=1).value[0:6] == "Пример":
            i += 1


            req = {'request': [{"role": "system", "text": txt_system},
                               {"role": "user", "text": txt_user}],
                   "response": txt_answer
                   }

            txt_user = "Проанализируй сообщение и распредели по группам\n Вот пример сообщения: "
            txt_answer = "\n"

            data.append(req)

            continue
        else:
            message = curr_sheet.cell(row=i, column=1).value
            txt_user += message
            i += 2

            continue




    for j in range(2, 10):
        if curr_sheet.cell(row=i, column=j).value == None:
            text = "  "
        else:
            text = str(curr_sheet.cell(row=i, column=j).value)

        txt_answer += col_names[j - 2] + ": " + text + '\n'

    txt_answer += '\n'
    i += 1

with open("./data/output.jsonl", "w", encoding="utf-8") as f:
    for item in data:
        f.write(json.dumps(item, ensure_ascii=False) + "\n")