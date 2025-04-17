from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.styles import Border, Side
import json
import logging

# Названия столбцов таблицы
col_names = ['Дата', 'Подразделение', 'Операция', 'Культура', 'За день, га', 'С начала операции, га', 'Вал за день, ц',
             'Вал с начала, ц']

# Настройка границ ячеек
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


def empty(curr_sheet, curr_row):
    """
            Проверяет, пустая ли строка в таблице.

            Args:
                :param curr_sheet: Текущий лист
                :param curr_row: Текущая строка
    """
    for i in range(0, 1):
        for j in range(1, 10):
            if curr_sheet.cell(row=curr_row + i, column=j).value != None:
                return False
    return True


def generate_table(table_name='model/data/table1.xlsx'):
    """
            Создает новую таблицу Excel с заголовками.

            Args:
                :param table_name: Название таблицы по умолачанию является тем, которое задано в ТЗ
    """
    try:
        wb = Workbook()
        wb.create_sheet('Лист1')

        # Удаляем лист по умолчанию 'Sheet', если он есть
        default_sheet = 'Sheet'
        if default_sheet in wb.sheetnames:
            sheet_to_remove = wb[default_sheet]
            wb.remove(sheet_to_remove)

        curr_sheet = wb['Лист1']

        # Заполняем заголовки столбцов
        for i in range(0, len(col_names)):
            curr_sheet.cell(row=2, column=i + 1, value=col_names[i])
            curr_sheet.cell(row=2, column=i + 1).font = Font(name='Calibri', size=12, bold=True)
            curr_sheet.cell(row=2, column=i + 1).fill = PatternFill(patternType='solid', start_color='C2D69B')
            curr_sheet.cell(row=2, column=i + 1).alignment = Alignment(horizontal='center')
            curr_sheet.cell(row=2, column=i + 1).border = thin_border

            # Настраиваем ширину столбцов
            if col_names[i] == 'Операция':
                curr_sheet.column_dimensions[chr(i + 65)].width = 35
            else:
                curr_sheet.column_dimensions[chr(i + 65)].width = 25

        wb.save(table_name)
    except Exception as e:
        logging.error(f"Ошибка при создании таблицы: {e}")
        raise


def write_data(file_name='model/processed_data/data.jsonl', table_name='model/data/table1.xlsx'):
    """
            Записывает данные из JSONL-файла в таблицу Excel.

            Args:
                :param file_name: Название JSONL-файла по умолачанию является тем, которое задается в assistant.py
                :param table_name: Название таблицы по умолачанию является тем, которое задано в ТЗ
    """
    try:
        with open(file_name, 'r', encoding='utf-8') as file:
            data = [json.loads(line) for line in file if line.strip()]

        wb = load_workbook(table_name)
        sheet_names = wb.sheetnames
        curr_sheet = wb[sheet_names[0]]  # Получаем объект листа, а не его имя
        i = 3  # Начинаем с 3 строки (после заголовков)

        for k in range(0, len(data)):
            # Пропускаем заполненные строки
            while curr_sheet.cell(row=i, column=2).value is not None:
                i += 1

            # Записываем данные в ячейки
            for j in range(0, len(col_names)):
                curr_sheet.cell(row=i, column=j + 1).value = data[k][col_names[j]]
                curr_sheet.cell(row=i, column=j + 1).font = Font(name='Calibri', size=12)
                curr_sheet.cell(row=i, column=j + 1).alignment = Alignment(horizontal='center')
                curr_sheet.cell(row=i, column=j + 1).border = thin_border

        wb.save(table_name)
    except FileNotFoundError:
        logging.error(f"Файл не найден: {file_name}")
        raise
    except json.JSONDecodeError:
        logging.error(f"Ошибка декодирования JSON в файле: {file_name}")
        raise
    except Exception as e:
        logging.error(f"Ошибка при записи данных: {e}")
        raise


def read_inf(curr_sheet, d):
    """
            Читает данные из справочного листа и добавляет их в список d.

            Args:
                :param sheet_name: Название листа
                :param d: Словарь, куда сохраняются данные
    """
    try:
        r = 4  # Начинаем с 4 строки

        while not empty(curr_sheet, r):
            value = curr_sheet.cell(row=r, column=1).value
            if value not in d:
                d.append(value)
            r += 1
    except Exception as e:
        logging.error(f"Ошибка при чтении справочной информации: {e}")
        raise


def check_table(table_name='model/data/table1.xlsx'):
    """
            Проверяет таблицу на соответствие справочным данным.

            Args:
                :param table_name: Название таблицы по умолачанию является тем, которое задано в ТЗ
    """
    try:
        # Загружаем справочные данные
        wb_ref = load_workbook('model/data/Справочная информация.xlsx')
        sheet_names_ref = wb_ref.sheetnames

        subdivision = []
        culture = []
        operation = []

        read_inf(wb_ref, wb_ref[sheet_names_ref[0]], subdivision)  # Подразделения
        read_inf(wb_ref, wb_ref[sheet_names_ref[1]], operation)  # Операции
        read_inf(wb_ref, wb_ref[sheet_names_ref[2]], culture)  # Культуры

        # Загружаем проверяемую таблицу
        wb = load_workbook(table_name)
        sheet_names = wb.sheetnames
        curr_sheet = wb[sheet_names[0]]  # Получаем объект листа
        r = 3  # Начинаем с 3 строки

        while not empty(curr_sheet, r):
            # Проверяем подразделение
            if curr_sheet.cell(row=r, column=2).value not in subdivision:
                curr_sheet.cell(row=r, column=2).fill = PatternFill(patternType='solid',
                                                                    start_color='00FFFF00')  # Желтый

            # Проверяем операцию
            if curr_sheet.cell(row=r, column=3).value not in operation:
                curr_sheet.cell(row=r, column=3).fill = PatternFill(patternType='solid', start_color='00FFFF00')

            # Проверяем культуру
            if curr_sheet.cell(row=r, column=4).value not in culture:
                curr_sheet.cell(row=r, column=4).fill = PatternFill(patternType='solid', start_color='00FFFF00')

            r += 1

        wb.save(table_name)
    except FileNotFoundError:
        logging.error("Справочный файл не найден!")
        raise
    except Exception as e:
        logging.error(f"Ошибка при проверке таблицы: {e}")
        raise