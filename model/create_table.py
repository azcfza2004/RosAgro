from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.styles import Border, Side
import json

col_names = ['Дата', 'Подразделение', 'Операция', 'Культура', 'За день, га', 'С начала операции, га', 'Вал за день, ц', 'Вал с начала, ц']
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

def generate_table():
    wb = Workbook()
    wb.create_sheet('Лист1')

    default_sheet = 'Sheet'
    if default_sheet in wb.sheetnames:
        sheet_to_remove = wb[default_sheet]
        wb.remove(sheet_to_remove)

    curr_sheet = wb['Лист1']

    for i in range(1, len(col_names) + 1):
        curr_sheet.cell(row=2, column=i+1, value=col_names[i - 1])
        curr_sheet.cell(row=2, column=i+1).font = Font(name='Calibri', size=12, bold=True)
        curr_sheet.cell(row=2, column=i+1).fill = PatternFill(patternType='solid', start_color='C2D69B')
        curr_sheet.cell(row=2, column=i+1).alignment = Alignment(horizontal='center')
        curr_sheet.cell(row=2, column=i+1).border = thin_border

        if col_names[i - 1] == 'Операция':
            curr_sheet.column_dimensions[chr(i + 65)].width = 35
        else:
            curr_sheet.column_dimensions[chr(i + 65)].width = 25

    wb.save('./data/table1_30.xlsx')

def write_data(last_id, file_name='res.jsonl'):
    with open('./data/' + file_name, 'r', encoding='utf-8') as file:
        # all_data = json.load(file)
        data = [json.loads(line) for line in file if line.strip()]

    # if last_id == all_data['metadata']['id']:
    #     return -1

    # data = all_data['data']

    last_id += 1

    wb = load_workbook('./data/table1_30.xlsx')

    curr_sheet = wb['Лист1']

    i = 2
    # print(len(data))
    for k in range(0, len(data)):
        print(k)
        while curr_sheet.cell(row=i, column=3).value != None or curr_sheet.cell(row=i, column=1).value != None:
            i += 1

        if len(data[k]) < 2:
            curr_sheet.cell(row=i, column=1).value = data[k]['message']
            # print(i, data[k])
            i += 1
            continue

        # print(i, data[k])
        for j in range(1, len(col_names) + 1):
            curr_sheet.cell(row=i, column=j+1).value = data[k][col_names[j - 1]]
            curr_sheet.cell(row=i, column=j+1).font = Font(name='Calibri', size=12)
            curr_sheet.cell(row=i, column=j+1).alignment = Alignment(horizontal='center')
            curr_sheet.cell(row=i, column=j+1).border = thin_border

    wb.save('./data/table1_30.xlsx')

generate_table()
write_data(-1)